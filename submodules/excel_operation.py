import datetime
from re import search
import openpyxl
import json

#
#record = [tmp_id, tmp_name, date, fever_status, cmp_avg, obj_avg, amb_avg]
#record = [tmp_id, tmp_name, date,               cmp_avg, obj_avg, amb_avg]
def search_for_last_row(worksheet):
    row = 4     # start from row 4
    col = 2     # Date
    while True:
        val = worksheet.cell(column=col, row=row).value
        if val == None or val=='':
            return row
        row += 1
        if row >= 50000:
            raise Exception('Too many records in the excel file (>50,000 rows)')
            break

def export_to_excel(filename:str, data:list):
    with open('thermometer_cfg.json', 'rt') as f:    # read configurations from the configuration file
        config = json.load(f)
    if config['system']['school_flag'] == 'True':
        dt = datetime.datetime.now()
        date_str = '{}/{}/{}'.format(dt.year, dt.month, dt.day)
        workbook = openpyxl.load_workbook(config['excel']['filename'])
        worksheet = workbook[config['excel']['worksheet']]

        row = search_for_last_row(worksheet)
        for line in data:
            student_id   = line[0]
            student_name = line[1]
            student_temp = line[4]
            print(student_id, student_name, student_temp)
            worksheet.cell(column=2, row=row, value=date_str)
            worksheet.cell(column=3, row=row, value=student_id)
            if student_temp < 37.0:
                worksheet.cell(column=4, row=row, value='37℃ 以下')
                worksheet.cell(column=5, row=row, value='')
            else:
                worksheet.cell(column=4, row=row, value='37℃ 以上')
                worksheet.cell(column=5, row=row, value=student_temp)
        workbook.save(config['excel']['filename'])
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for row, line in enumerate(data):
            for col, dt in enumerate(line):
                worksheet.cell(column=col+1, row=row+1, value=dt)
        workbook.save(filename)
